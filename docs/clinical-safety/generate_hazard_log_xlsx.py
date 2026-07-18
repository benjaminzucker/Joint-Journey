#!/usr/bin/env python3
"""
Generate: Joint Journey — Hazard Log (DCB0129) as a formatted Excel workbook.

Produces a colour-coded .xlsx with three sheets:
  1. Hazard Log  — all hazards with risk scores
  2. Open Actions — outstanding items before go-live
  3. Change Log   — version history

Run:
    python3 generate_hazard_log_xlsx.py

Requires: openpyxl  (pip3 install openpyxl)
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "Joint Journey - Hazard Log (DCB0129).xlsx")

# ── Colours ──────────────────────────────────────────────────────────────
NAVY = "1F3A5F"
WHITE = "FFFFFF"
LIGHT_GREEN = "E8F5E9"
LIGHT_AMBER = "FFF8E1"
LIGHT_RED = "FFEBEE"
LIGHT_GREY = "F5F5F5"
GREEN = "4CAF50"
AMBER = "FF9800"
RED = "F44336"
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)
META_FONT = Font(name="Calibri", size=11, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def risk_fill(score):
    """Return a fill colour based on residual risk score."""
    if score <= 4:
        return PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
    elif score <= 6:
        return PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid")
    else:
        return PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")


def risk_font(score):
    """Return a font colour based on risk score."""
    if score <= 4:
        return Font(name="Calibri", size=11, bold=True, color=GREEN)
    elif score <= 6:
        return Font(name="Calibri", size=11, bold=True, color=AMBER)
    else:
        return Font(name="Calibri", size=11, bold=True, color=RED)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER


def style_body_cell(ws, row, col, font=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font or BODY_FONT
    cell.alignment = WRAP
    cell.border = THIN_BORDER
    return cell


# ── Hazard data ──────────────────────────────────────────────────────────
HAZARDS = [
    {
        "id": "H01",
        "title": "User performs an exercise that is unsafe for them and is injured",
        "cause": "Generic programme not suited to a specific comorbidity; user pushes too hard.",
        "effect": "Musculoskeletal injury, fall, pain flare.",
        "init_l": 3, "init_s": 3,
        "controls": (
            "\"Stop if sharp pain\" guidance; \"hold onto something sturdy\"; "
            "\"talk to your GP before starting\"; gentle/standard/active levelling; "
            "warm-up instructions. Pre-exercise safety check collapsible (v0.2). "
            "PAR-Q-style safety screening questionnaire during onboarding (v0.4): "
            "heart/lung, balance/falls, GP-restricted exercise. Flagged users see "
            "persistent caution banner on exercise page."
        ),
        "further": "Video demonstrations of correct form.",
        "res_l": 2, "res_s": 2,
        "status": "Implemented (v0.4) — screening questionnaire + pre-exercise safety check live.",
        "domain": "Physiotherapy",
    },

    {
        "id": "H02",
        "title": "User relies on the app instead of seeking medical help; a problem is missed",
        "cause": "User interprets app as a substitute for clinical care.",
        "effect": "Delayed diagnosis/treatment (e.g. DVT, infection, worsening condition).",
        "init_l": 2, "init_s": 4,
        "controls": (
            "\"Not a substitute for medical advice\" disclaimers. Persistent \"When to seek help\" "
            "control in app header on every in-app view — panel lists 999 / 111 / GP / surgical team (v0.2)."
        ),
        "further": "Extend safety-netting copy into education/\"after surgery\" modules as content grows.",
        "res_l": 1, "res_s": 4,
        "status": "Implemented (v0.2).",
        "domain": "CSO / multidisciplinary",
    },
    {
        "id": "H03",
        "title": "Oxford-Score banding places user in the wrong programme level",
        "cause": "User mis-enters score; rule edge cases; \"both joints\" logic.",
        "effect": "Programme too hard (injury risk) or too easy (suboptimal prep).",
        "init_l": 3, "init_s": 2,
        "controls": "Validated Oxford questionnaire; \"both joints\" uses lower score; user can self-adjust level.",
        "further": "Allow easy level change; \"this felt too hard/easy?\" prompt; boundary checks.",
        "res_l": 2, "res_s": 2,
        "status": "Controls in place.",
        "domain": "Physiotherapy",
    },
    {
        "id": "H04",
        "title": "Incorrect nutrition / calorie information leads to harmful dieting",
        "cause": "Generic calorie/BMI calculation applied to someone for whom it's inappropriate.",
        "effect": "Unsafe weight loss / nutritional harm.",
        "init_l": 2, "init_s": 3,
        "controls": (
            "Presented as general guidance; estimates noted as approximate. "
            "Deficit only applied when BMI indicates weight to lose; capped at safe maximum "
            "(300–500 kcal/day); never below absolute calorie floor. "
            "Collapsible \"Consult your doctor\" safety check on nutrition page (v0.4): "
            "diabetes, eating disorders, kidney disease, heart failure, warfarin, underweight, pregnancy."
        ),
        "further": "None outstanding — core controls in place.",
        "res_l": 1, "res_s": 3,
        "status": "Implemented (v0.4) — deficit cap + condition cautions live.",
        "domain": "Dietetics",
    },

    {
        "id": "H05",
        "title": "Mental-health content distresses a vulnerable user / misses risk",
        "cause": "Modules on pain/anxiety surface distress; no route for a user in crisis.",
        "effect": "Worsening distress; missed safeguarding/crisis situation.",
        "init_l": 2, "init_s": 4,
        "controls": (
            "Evidence-based supportive content. Crisis signposting — persistent \"When to seek help\" "
            "panel surfaces Samaritans 116 123, 111/999, GP routes (v0.2)."
        ),
        "further": "Reinforce \"this is not therapy\" framing; repeat crisis signposting within individual mindset modules.",
        "res_l": 1, "res_s": 4,
        "status": "Implemented (v0.2) — crisis signposting live site-wide.",
        "domain": "Psychology",
    },
    {
        "id": "H06",
        "title": "Software error displays wrong data (weight, score, progress)",
        "cause": "Bug; sync/caching issue; data-write failure.",
        "effect": "User confusion; wrong self-management decision.",
        "init_l": 2, "init_s": 2,
        "controls": "Data shown is user's own; no clinical decision driven by it.",
        "further": "Input validation; testing; release checks.",
        "res_l": 1, "res_s": 2,
        "status": "Controls in place.",
        "domain": "CSO / multidisciplinary",
    },
    {
        "id": "H07",
        "title": "Data loss / breach of personal health data",
        "cause": "Security/hosting failure; misconfiguration.",
        "effect": "Privacy harm; loss of trust.",
        "init_l": 2, "init_s": 3,
        "controls": (
            "Authenticated accounts; consent at signup; Firestore in europe-west2. "
            "Firestore security rules audited (PASS). Cross-ref: DPIA (docs/dpia.md)."
        ),
        "further": "Cyber Essentials certification; automated backups; incident-response procedure (72h ICO).",
        "res_l": 1, "res_s": 3,
        "status": "DPIA cross-referenced (v0.3). Cyber Essentials still to complete.",
        "domain": "CSO / multidisciplinary",
    },
    {
        "id": "H08",
        "title": "Accessibility barrier excludes / misleads older or impaired users",
        "cause": "Poor contrast, small targets, no captions → misreads instructions.",
        "effect": "Misuse of exercises; exclusion.",
        "init_l": 3, "init_s": 2,
        "controls": "Large fonts, simple layout.",
        "further": "WCAG 2.1 AA review; captions on videos; tested with older users.",
        "res_l": 2, "res_s": 2,
        "status": "OPEN — WCAG pass still to complete.",
        "domain": "CSO / multidisciplinary",
    },
    {
        "id": "H09",
        "title": "User follows outdated or superseded clinical content",
        "cause": "Guidelines change after content was written; content not reviewed/updated.",
        "effect": "User follows advice that is no longer current best practice.",
        "init_l": 2, "init_s": 2,
        "controls": "Content developed with reference to current NICE/CSP guidelines (June–July 2026). Version-controlled in git.",
        "further": "Scheduled annual content review; \"last reviewed\" dates visible to users; guideline monitoring.",
        "res_l": 1, "res_s": 2,
        "status": "Low risk for new product. Added v0.3.",
        "domain": "CSO / advisory board",
    },
    {
        "id": "H10",
        "title": "Third-party service failure disrupts the user experience",
        "cause": "Firebase outage; YouTube/Vimeo video removed; hosting (Netlify) down.",
        "effect": "User cannot access exercises, content, or data; interruption to programme.",
        "init_l": 2, "init_s": 2,
        "controls": "Firebase/Netlify high-availability SLAs (99.95%+). Local-storage caching. Text-based exercise descriptions.",
        "further": "PWA offline mode (service worker); self-hosted video fallback; uptime monitoring.",
        "res_l": 1, "res_s": 2,
        "status": "Low risk. No clinical harm from temporary unavailability. Added v0.3.",
        "domain": "CSO / multidisciplinary",
    },
]

OPEN_ACTIONS = [
    ("CSO appointed + trained.", "DONE", "v0.3 — Mr Benjamin Zucker registered as CSO."),
    ("Red-flag / \"when to seek help\" content (H02, H05).", "DONE", "v0.2 — persistent panel on every in-app view."),
    ("Crisis signposting including Samaritans (H05).", "DONE", "v0.2 — in the persistent panel site-wide."),
    ("Pre-exercise safety screening (H01).", "DONE", "v0.4 — PAR-Q-style screening questionnaire in onboarding + caution banner on exercise page. Pre-exercise safety check also live (v0.2)."),
    ("DPIA cross-referenced for H07.", "DONE", "v0.3 — DPIA path, data controller, data residency added."),
    ("Nutrition safety controls (H04).", "DONE", "v0.4 — Deficit capped at 300–500 kcal/day, only applied when BMI indicates weight to lose. \"Consult your doctor\" collapsible on nutrition page."),
    ("WCAG 2.1 AA accessibility pass (H08).", "OPEN", "Check contrast, touch targets, alt text, keyboard nav."),
    ("Clinical advisory board review and sign-off (H01–H10).", "OPEN", "Schedule first meeting; record in content-review-and-signoff-log."),
    ("Formal CSO sign-off of this log (bump to v1.0).", "OPEN", "After all above are closed."),
]

CHANGES = [
    ("0.4", "18/07/2026", (
        "Safety screening questionnaire added to onboarding (PAR-Q-style: heart/lung, "
        "balance/falls, exercise restriction). Caution banner on exercise page for flagged users. "
        "\"Consult your doctor\" collapsible added to nutrition page. "
        "H01 residual risk reduced to 2×2=4. H04 controls updated, action closed."
    )),
    ("0.3", "18/07/2026", (
        "CSO registered (action closed). Manufacturer updated to Elan Health Ltd (17347255). "
        "DPIA cross-reference added to H07 (action closed). H09 and H10 added. "
        "Excel hazard log generator created."
    )),

    ("0.2", "13/07/2026", (
        "In-app safety controls implemented: persistent \"When to seek help\" header control "
        "(H02), crisis signposting including Samaritans (H05), pre-exercise safety check (H01)."
    )),
    ("0.1", "29/06/2026", "Initial draft. H01–H08 identified. First-pass risk ratings."),
]


def build():
    wb = Workbook()

    # =====================================================================
    # SHEET 1 — HAZARD LOG
    # =====================================================================
    ws = wb.active
    ws.title = "Hazard Log"
    ws.sheet_properties.tabColor = NAVY

    # Title block
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "Joint Journey — Clinical Safety Hazard Log (DCB0129)"
    c.font = TITLE_FONT

    meta_lines = [
        ("A2", "Standard: DCB0129 (manufacturer of health IT)"),
        ("A3", "Manufacturer: Elan Health Ltd (company number 17347255)"),
        ("A4", "Clinical Safety Officer: Mr Benjamin Zucker (registered CSO)"),
        ("A5", "Version: 0.3 (draft)  |  Date: 18/07/2026  |  Status: Working draft — CSO registered; awaiting advisory board review"),
    ]
    for ref, text in meta_lines:
        ws[ref].value = text
        ws[ref].font = META_FONT

    # Column headers (row 7)
    headers = [
        "ID", "Hazard", "Cause", "Effect",
        "Init L", "Init S", "Init Risk",
        "Existing Controls",
        "Further Mitigations",
        "Res L", "Res S", "Residual Risk",
        "Status", "Domain / Reviewer"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=h)
    style_header_row(ws, 7, len(headers))

    # Data rows
    for i, h in enumerate(HAZARDS):
        row = 8 + i
        init_risk = h["init_l"] * h["init_s"]
        res_risk = h["res_l"] * h["res_s"]
        values = [
            h["id"], h["title"], h["cause"], h["effect"],
            h["init_l"], h["init_s"], init_risk,
            h["controls"], h["further"],
            h["res_l"], h["res_s"], res_risk,
            h["status"], h["domain"],
        ]
        for col, v in enumerate(values, 1):
            cell = style_body_cell(ws, row, col)
            cell.value = v

        # Colour-code initial risk
        ws.cell(row=row, column=7).fill = risk_fill(init_risk)
        ws.cell(row=row, column=7).font = risk_font(init_risk)

        # Colour-code residual risk
        ws.cell(row=row, column=12).fill = risk_fill(res_risk)
        ws.cell(row=row, column=12).font = risk_font(res_risk)

        # Alternate row shading
        if i % 2 == 1:
            grey = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                if col not in (7, 12):  # don't overwrite risk colours
                    cell.fill = grey

    # Column widths
    widths = [5, 40, 35, 30, 6, 6, 8, 50, 40, 6, 6, 10, 35, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze panes
    ws.freeze_panes = "A8"

    # Risk matrix key (below hazards)
    key_row = 8 + len(HAZARDS) + 2
    ws.cell(row=key_row, column=1, value="Risk scoring key:").font = BOLD_FONT
    ws.cell(row=key_row + 1, column=1, value="≤ 4 = Acceptable (green)").font = Font(name="Calibri", size=10, color=GREEN)
    ws.cell(row=key_row + 2, column=1, value="5–6 = Tolerable (amber)").font = Font(name="Calibri", size=10, color=AMBER)
    ws.cell(row=key_row + 3, column=1, value="≥ 7 = Unacceptable (red)").font = Font(name="Calibri", size=10, color=RED)
    ws.cell(row=key_row + 4, column=1, value="Likelihood: 1 Very low · 2 Low · 3 Medium · 4 High · 5 Very high").font = META_FONT
    ws.cell(row=key_row + 5, column=1, value="Severity: 1 Minor · 2 Significant · 3 Serious · 4 Major · 5 Catastrophic").font = META_FONT

    # =====================================================================
    # SHEET 2 — OPEN ACTIONS
    # =====================================================================
    ws2 = wb.create_sheet("Open Actions")
    ws2.sheet_properties.tabColor = "FF9800"

    ws2.merge_cells("A1:C1")
    ws2["A1"].value = "Open Actions — to close before deployment"
    ws2["A1"].font = TITLE_FONT

    action_headers = ["Action", "Status", "Notes"]
    for col, h in enumerate(action_headers, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header_row(ws2, 3, len(action_headers))

    status_fills = {
        "DONE": PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid"),
        "PARTIAL": PatternFill(start_color=LIGHT_AMBER, end_color=LIGHT_AMBER, fill_type="solid"),
        "OPEN": PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid"),
    }

    for i, (action, status, notes) in enumerate(OPEN_ACTIONS):
        row = 4 + i
        style_body_cell(ws2, row, 1).value = action
        cell_s = style_body_cell(ws2, row, 2)
        cell_s.value = status
        cell_s.font = BOLD_FONT
        cell_s.fill = status_fills.get(status, PatternFill())
        style_body_cell(ws2, row, 3).value = notes

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 60
    ws2.freeze_panes = "A4"

    # =====================================================================
    # SHEET 3 — CHANGE LOG
    # =====================================================================
    ws3 = wb.create_sheet("Change Log")
    ws3.sheet_properties.tabColor = "4CAF50"

    ws3.merge_cells("A1:C1")
    ws3["A1"].value = "Change Log"
    ws3["A1"].font = TITLE_FONT

    cl_headers = ["Version", "Date", "Changes"]
    for col, h in enumerate(cl_headers, 1):
        ws3.cell(row=3, column=col, value=h)
    style_header_row(ws3, 3, len(cl_headers))

    for i, (ver, date, desc) in enumerate(CHANGES):
        row = 4 + i
        style_body_cell(ws3, row, 1).value = ver
        style_body_cell(ws3, row, 2).value = date
        style_body_cell(ws3, row, 3).value = desc

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 90
    ws3.freeze_panes = "A4"

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    build()
